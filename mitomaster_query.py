#!/usr/bin/env python3
"""Submit mitochondrial consensus FASTA to MITOMASTER web API and save as Excel."""

import io
import os

import pandas as pd
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_URL = "https://www.mitomap.org/mitomaster/websrvc.cgi"
API_KEY = "1776426752482"

FASTA_FILE = os.path.join(os.path.dirname(__file__), "your_patient.mt.consensus.fasta")


def query_mitomaster(fasta_path: str) -> str:
    with open(fasta_path, "rb") as f:
        response = requests.post(
            API_URL,
            data={"key": API_KEY, "fileType": "sequences", "output": "detail"},
            files={"file": (os.path.basename(fasta_path), f, "text/plain")},
            timeout=120,
        )
    response.raise_for_status()
    return response.text


def save_excel(df: pd.DataFrame, sample_name: str, out_path: str) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Write data starting at row 2 (row 1 = title)
        df.to_excel(writer, index=False, sheet_name="Sheet1", startrow=1)
        ws = writer.sheets["Sheet1"]

        # Row 1: title
        title_cell = ws.cell(row=1, column=1, value=f"Mitomaster {sample_name}")
        title_cell.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

        # Header row styling (row 2)
        header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, _ in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        # Data rows styling + column widths
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[2].height = 40
        ws.freeze_panes = "A3"


COLUMN_MAP = {
    "query":          "Query",
    "tpos":           "rCRS Position",
    "tnt":            "rCRS NT",
    "qnt":            "Query NT",
    "ntchange":       "Alt name†",
    "allele":         "Mut type",
    "calc_locus":     "Locus",
    "calc_aachange":  "Other",
    "gb_cnt":         "GB SeqsCount",
    "gb_perc":        "GB Freq %",
    "hap_cnt":        "Haplogroup Count",
    "hap_perc":       "Freq % in W4d",
    "conservation":   "Conservation",
    "patientphenotype": "Patient Report◊",
}

COLUMN_ORDER = list(COLUMN_MAP.values())


def reshape(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns=COLUMN_MAP)
    existing = [c for c in COLUMN_ORDER if c in df.columns]
    return df[existing]


def main() -> None:
    sample_name = os.path.basename(FASTA_FILE).replace(".mt.consensus.fasta", "")
    print(f"Submitting {os.path.basename(FASTA_FILE)} to MITOMASTER...")

    result = query_mitomaster(FASTA_FILE)
    df = reshape(pd.read_csv(io.StringIO(result), sep="\t"))

    out_file = os.path.join(os.path.dirname(FASTA_FILE), f"Mitomaster {sample_name}.xlsx")
    save_excel(df, sample_name, out_file)

    print(f"Done. Output saved to: {os.path.basename(out_file)}")
    print(f"  {len(df)} variants  |  {len(df.columns)} columns")


if __name__ == "__main__":
    main()
